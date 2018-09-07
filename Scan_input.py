# -*- coding: utf-8 -*-
"""
Created on Fri Jul 13 10:42:38 2018

@author: ukn1hc
"""

import sys

filepath = sys.argv[1]
byteoder = sys.argv[2] # 'm' = msb or 'i' = intel


# Load file Excel input
from openpyxl import load_workbook
wb = load_workbook(filename = filepath)
sheet_ranges = wb['Sheet1'] # grab the active worksheet


# Prepare the excel output
from openpyxl import Workbook
wb1 = Workbook()
ws = wb1.active # grab the active worksheet

excel_title=['ID',
             "Frame Name",
             "Cycle Time [ms]",
             "Launch Type",
             "Launch Parameter",
             "Signal Byte No.",
             "Signal Bit No.",
             "Signal Name",
             "Signal Function",
             "Signal Length [Bit]",
             "Signal Default",
             "Signal Not Available",
             "Byteorder",
             "Nodes",
             "Vector__XXX",
             "Value",
             "Name / Phys. Range",
             "Function / Increment Unit"]

# Add title row
for i in range(0,len(excel_title)):
    ws.cell(row = 1, column = i+1, value = excel_title[i])
    


wb1.save("excel2dbc_{}".format(filepath))

