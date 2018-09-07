# -*- coding: utf-8 -*-
"""
Created on Fri Jul  6 11:15:40 2018

@author: ukn1hc
"""

# -----------------------------------------------------------------------------------
# http://openpyxl.readthedocs.io/en/stable
# -----------------------------------------------------------------------------------

# -----------------------------------------------------------------------------------
# Create new excel file
# -----------------------------------------------------------------------------------


# from openpyxl import Workbook
# wb = Workbook()

# # grab the active worksheet
# ws = wb.active

# # Data can be assigned directly to cells
# ws['A1'] = 42

# # Rows can also be appended
# ws.append([1, 2, 3])

# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()

# # Save the file
# wb.save("sample.xlsx")

# -----------------------------------------------------------------------------------
# Loading from a file
# -----------------------------------------------------------------------------------
# from openpyxl import load_workbook
# wb2 = load_workbook('sample.xlsx')


# -----------------------------------------------------------------------------------
# Write a workbook
# -----------------------------------------------------------------------------------

# from openpyxl import Workbook
# from openpyxl.compat import range
# from openpyxl.utils import get_column_letter
# wb = Workbook()

# dest_filename = 'empty_book.xlsx' 

# ws1 = wb.active
# ws1.title = "range names" #Name of the sheet 1
# for row in range(1, 40):
# 	ws1.append(range(600))
# ws2 = wb.create_sheet(title="Pi") # Create sheet 2 : Pi
# ws2['F5'] = 3.14
# ws3 = wb.create_sheet(title="Data") # Create sheet 23 : DataPi
# for row in range(10, 20):
# 	for col in range(27, 54):
# 		_ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
# print(ws3['AA10'].value)
# #AA

# wb.save(filename = dest_filename)

# -----------------------------------------------------------------------------------
# Read an existing workbook
# -----------------------------------------------------------------------------------

from openpyxl import load_workbook
wb = load_workbook(filename = 'empty_book.xlsx')
sheet_ranges = wb['range names']  # Load the sheet name. In usual: Sheet1. We can call ws = wb['range names']
print(sheet_ranges['D18'].value)
# 3
print(sheet_ranges.cell(3,17).value) #(row , col)
# 16



#------------------------------
# read current row and column
# cell.row and cell.column
#----
#or 
#sheet_ranges['{}{}'.format(excel_title['ID'],cell.row)].value
#------------------------------

# -----------------------------------------------------------------------------------
# Iterate through all rows in specific column openpyxl
# -----------------------------------------------------------------------------------
# ws.iter_rows() or sheet_ranges
# ws.iter_cols()
# ws[E] and ws[E:Q] or ws[4] and ws[4:6]
# -----------------------------------------------------------------------------------


# Read all valid value row in 'C' col
# for cell in sheet_ranges['C']:
#    print(cell.value) 

# Read all valid value col in 4 row. Openpixl use index 1.
for cell in sheet_ranges[1]:
   print(cell.value) 


# Use iter_row or col. But deprecated
# for row in sheet_ranges.iter_rows('C{}:C{}'.format(sheet_ranges.min_row,sheet_ranges.max_row)):
# 	for cell in row:
#         print(cell.value) 


# Scan a database . deprecated
# sheet_ranges.get_squared_range(min_col=1, min_row=1, max_col=1, max_row=10)


# A range of min_row to max_row - from col 4 to col 6.
# print(sheet_ranges[4:6])
# print(sheet_ranges[4])


# Add to list
# mylist = []
# for row in ws.iter_rows('A{}:A{}'.format(ws.min_row,ws.max_row)):
#     for cell in row:
#         mylist.append(cell.value)
# print mylist


# -----------------------------------------------------------------------------------
# Using number formats
# -----------------------------------------------------------------------------------
# import datetime
# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active
# # set date using a Python datetime
# ws['A1'] = datetime.datetime(2010, 7, 21)
# ws['A1'].number_format
# #'yyyy-mm-dd h:mm:ss'
# # You can enable type inference on a case-by-case basis
# wb.guess_types = True
# # set percentage using a string followed by the percent sign
# ws['B1'] = '3.14%'
# wb.guess_types = False
# ws['B1'].value
# # 0.031400000000000004
# ws['B1'].number_format
# # '0%'


# -----------------------------------------------------------------------------------
# Using formulae
# -----------------------------------------------------------------------------------
# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active
# # add a simple formula
# ws["A1"] = "=SUM(1, 1)"
# wb.save("formula.xlsx")

# -----------------------------------------------------------------------------------
# Fold columns (outline)
# -----------------------------------------------------------------------------------
# import openpyxl
# wb = openpyxl.Workbook()
# ws = wb.create_sheet()
# ws.column_dimensions.group('A','D', hidden=True)
# wb.save('group.xlsx')